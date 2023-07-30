import re
import requests
import openpyxl
from bs4 import BeautifulSoup

from checkurl import geturl

def scraping(url):
    try:
        response = requests.get(url)

        if response.status_code == 200:
            soup = BeautifulSoup(response.content, "html.parser")
            article_element = soup.find("article")

            if article_element:
                return article_element.prettify()
            else:
                print("tidak ada tag article : ", url)
                return None
        else:
            print("Gagal get content web. status:", response.status_code)
            return None

    except requests.exceptions.RequestException as e:
        print("Error : ", e)
        return None

def imageregx(html_content):
    soup = BeautifulSoup(html_content, "html.parser")
    div_element = soup.find("div", class_="bg-cover 3xl:bg-auto bg-center bg-no-repeat relative")
    if div_element:
        style_attr = div_element.get("style")
        if style_attr:
            match = re.search(r"url\('([^']+)'\)", style_attr)
            if match:
                return match.group(1)

    return None

def excelsave(contents, file_name):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for idx, content in enumerate(contents, start=1):
        soup = BeautifulSoup(content, "html.parser")

        # Find the specific div element and extract its text
        div_element = soup.find("div", class_="sm:text-sm md:text-2xl lg:text-5xl font-bold")
        if div_element:
            title_text = div_element.text.strip()
        else:
            title_text = "tidak ada teks"

        image_url = imageregx(content)

        sheet.cell(row=idx, column=1, value=content)
        sheet.cell(row=idx, column=2, value=title_text)
        sheet.cell(row=idx, column=3, value=image_url)
        print(f"judul {title_text}", f"gambar : {image_url}")

    workbook.save(file_name)
    print("Data berhasil disimpan : ", file_name)

urls = geturl()
print(urls)

isi_content = []
for url in urls:
    content = scraping(url)
    if content:
        isi_content.append(content)

if isi_content:
    excelsave(isi_content, "coba lgi wawawa.xlsx")
else:
    print("Gagal scraping.")
